"""
Everything about this program you can read on my Github
https://github.com/ToxicBlades/Mail_Parser_with_chatgpt
"""
import imaplib
import base64
import os
import openpyxl
import xlrd
import mailparser
from bs4 import BeautifulSoup
import openai
import logging
import time
import re
from pdfminer.high_level import extract_text
import mysql.connector
from docx import Document
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import csv
from sshtunnel import SSHTunnelForwarder
import pymysql
import io

#Secure this data :)
from config import EMAIL, PASSWORD, DIR , APIKEY ,SSH_PASSWORD, SSH_USERNAME, DATABASE_PASSWORD, DATABASE_USERNAME, SSH_HOST, DBNAME


ssh_host = SSH_HOST
ssh_username = SSH_USERNAME
ssh_password = SSH_PASSWORD
database_username = DATABASE_USERNAME
database_password = DATABASE_PASSWORD
database_name = DBNAME
localhost = '127.0.0.1'

tunnel = SSHTunnelForwarder(
        (ssh_host, 22),
        ssh_username = ssh_username,
        ssh_password = ssh_password,
        remote_bind_address = ('127.0.0.1', 3306)
    )

tunnel.start()

connection = mysql.connector.connect(
        host='127.0.0.1',
        user=database_username,
        passwd=database_password,
        db=database_name,
        port=tunnel.local_bind_port
    )

cursor = connection.cursor()


#pattern for checking messages
pattern = r"\s-\s(?![\w\d]+-[а-яА-ЯёЁ]+)"





#Set your apikey for chat gpt
openai.api_key = APIKEY

#Where are we gonna save our file after work
OUTPUT_DIR = DIR

def remove_text_and_words(input_string):
    words_to_remove = ["Forwarded", "message","messages", "Date", "Subject", "To", "Cc", "www","Adress","Mail","email","e-mail","Cell","Skype","Whatsapp","Warehouse","Manager","https","Copyright","Director","Web","Mobile","Phone","Mob","Ustr","Address","Tel","web-chat","T:","F:","regards","http","Trade"]
    russian_pattern = re.compile("[а-яА-ЯёЁ]+")  # find Russian words

    output_string = input_string
    for word in words_to_remove:
        pattern_n = r"(?i)\b" + word + r"\b.*?[\r\n]"  # case-insensitive match for the whole word until newline or carriage return
        output_string = re.sub(pattern_n, "", output_string)

    output_string = re.sub(russian_pattern, "", output_string)  # delete Russian words
    output_string = output_string.replace('\r', '')  # remove carriage return characters

    return output_string


def clear_string_for_db(input_string):
    words_to_clear_ai = [ 'Name','Product', 'Weight/Volume', 'Price', 'type of packaging','Type of packaging','type of price' , 'Weight/volume', 'of price' ,':',' : ','Incoterms','Incoterm','Type']
    for word in words_to_clear_ai:
        input_string = input_string.replace(word, '')
    return input_string.strip()


def connect_to_gmail(username, password):
    """Connects to Gmail IMAP server"""
    M = imaplib.IMAP4_SSL('imap.gmail.com')
    M.login(username, password)
    M.select('inbox')
    return M


def fetch_all_emails(M):
    """Fetches all emails from the inbox"""
    #Change UNSEEN TO ALL for first time if needed
    result, data = M.search(None, 'UNSEEN')
    email_ids = data[0].split()
    return email_ids


def send_email(subject, sender, date, to_email):
    """Send email using a Gmail account"""

    from_email = EMAIL  # your email here
    from_password = PASSWORD  # your password for apps here

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = "Этот офер нужно обработать вручную"

        body = f"Subject: {subject}\nSender: {sender}\nDate: {date}"
        msg.attach(MIMEText(body, 'plain'))
        time.sleep(20) #do not spam if a lot messages in a row :)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, from_password)
        text = msg.as_string()
        server.sendmail(from_email, to_email, text)
        server.quit()

    except Exception as e:
        logging.error("This email havent been send for some reason ")


def extract_email_fields(mail):
    """Extracts relevant fields from the email"""
    subject = str(mail.subject)
    sender = ''.join(map(str, mail.from_))
    date = str(mail.date)
    text = ''.join(map(str, mail.text_plain))
    text_html = ''
    #this thing exist for cases when for some reason we cant recive plain text from mail, so we get info with html
    if len(text) < 5:
        code_html = str(mail.text_html)
        soup = BeautifulSoup(code_html, 'html.parser')
        text_html = soup.get_text(separator=' ').strip()
    return subject, sender, date, text, text_html


def process_attachments(mail, attached_data):
    """Processes attachments in the email"""
    for attachment in mail.attachments:
        filename = attachment['filename']
        extension = os.path.splitext(filename)[1].lower()

        if extension == '.xlsx':
            file_data = attachment['payload']
            file_content = base64.b64decode(file_data)
            temp_filename = os.path.join(OUTPUT_DIR, filename)
            with open(temp_filename, 'wb') as temp_file:
                temp_file.write(file_content)
            xls_data = openpyxl.load_workbook(temp_filename)
            for sheet_name in xls_data.sheetnames:
                xls_sheet = xls_data[sheet_name]
                for row in xls_sheet.iter_rows(values_only=True):
                    row_data = '|'.join(map(str, row))
                    attached_data += row_data + '\n'
            os.remove(temp_filename)

        elif extension == '.xls':
            file_data = attachment['payload']
            file_content = base64.b64decode(file_data)
            temp_filename = os.path.join(OUTPUT_DIR, filename)
            with open(temp_filename, 'wb') as temp_file:
                temp_file.write(file_content)
            xls_data = xlrd.open_workbook(temp_filename)
            for sheet_name in xls_data.sheet_names():
                xls_sheet = xls_data.sheet_by_name(sheet_name)
                for row_index in range(xls_sheet.nrows):
                    row_data = '|'.join(map(str, xls_sheet.row_values(row_index)))
                    attached_data += row_data + '\n'
            os.remove(temp_filename)

        elif extension == '.pdf':
            file_data = attachment['payload']
            file_content = base64.b64decode(file_data)
            temp_filename = os.path.join(OUTPUT_DIR, filename)
            with open(temp_filename, 'wb') as temp_file:
                temp_file.write(file_content)

            pdf_text = extract_text(temp_filename)
            pdf_text = pdf_text.replace("\f","")
            attached_data += pdf_text
            os.remove(temp_filename)

        elif extension == '.docx':
            file_data = attachment['payload']
            file_content = base64.b64decode(file_data)
            temp_filename = os.path.join(OUTPUT_DIR, filename)
            with open(temp_filename, 'wb') as temp_file:
                temp_file.write(file_content)

            docx_text = ''
            doc = Document(temp_filename)
            for para in doc.paragraphs:
                docx_text += para.text + '\n'
            attached_data += docx_text

            os.remove(temp_filename)

        elif extension == '.csv':
            file_data = attachment['payload']
            file_content = base64.b64decode(file_data)
            temp_filename = os.path.join(OUTPUT_DIR, filename)
            with open(temp_filename, 'wb') as temp_file:
                temp_file.write(file_content)
            with open(temp_filename, 'r') as csv_file:
                csv_reader = csv.reader(csv_file)
                for row in csv_reader:
                    row_data = '|'.join(map(str, row))
                    attached_data += row_data + '\n'
            os.remove(temp_filename)

    return attached_data


def parse_emails(username, password):
    """Main function to parse emails"""
    try:
        M = connect_to_gmail(username, password)
        email_ids = fetch_all_emails(M)

        email_data = []
        for email_id in email_ids:
            try:
                email_data.append(process_email(M, email_id))
            except Exception as e:
                logging.error(f"Failed to process email {email_id}: {str(e)}")
    except Exception as e:
        logging.error(f"Failed to connect or fetch emails: {str(e)}")
        return

    process_ai(email_data)

    M.close()
    M.logout()


def process_email(M, email_id):
    """Function to process a single email"""
    result, data = M.fetch(email_id, '(RFC822)')
    mail = mailparser.parse_from_bytes(data[0][1])

    subject, sender, date, text, text_html = extract_email_fields(mail)

    attached_data = process_attachments(mail, '')

    if text_html:
        return [subject, sender, date, text, text_html, attached_data,mail]
    else:
        return [subject, sender, date, text, '', attached_data,mail]


ai_responses = []  # List to save AI responses


def process_ai(email_data):
    """Function to perform AI processing on email data"""
    for data in email_data:
        subject = data[0]
        text = data[3]
        mail = data[6]
        if "Fwd:" in subject:
            sender = ""
            date = ""
            subject = subject.replace("Fwd:","") #delete useless info
            lines = text.split("\n")  # Split the email text into lines

            for line in lines:
                if line.startswith("From:") or line.startswith("От:"):
                    sender = line.replace("From:", "").replace("От:", "").strip()
                    sender = sender.replace("<","").replace(">","")
                elif line.startswith("Date:"):
                    date = line.replace("Date:", "").strip()
        else:
            sender = data[1]
            sender = sender.replace("<","").replace(">","")
            date = data[2]

        text_html = data[4]
        attached_data = data[5]
        response_data = ''
        if attached_data:
            if len(attached_data)<15000:
                if text_html:
                    combined_text = remove_text_and_words(text_html + "\n" + attached_data)
                else:
                    combined_text = remove_text_and_words(text + "\n" + attached_data)
            else:
                combined_text = remove_text_and_words(text)
                for attachment in mail.attachments:
                    filename = attachment['filename']
                    if filename.endswith(('.xls', '.xlsx', '.pdf', '.csv')):
                        sql = "INSERT INTO offers (product,weight, pack_type, price, price_type, incoterm,sender,subject,date) VALUES (%s, %s, %s, %s,%s,%s,%s,%s,%s)"
                        product = filename
                        weight = 'none'
                        pack_type = 'none'
                        price = 'none'
                        price_type = 'none'
                        incoterm = 'none'
                        values = (product,weight, pack_type, price, price_type, incoterm,sender,subject,date)
                        cursor.execute(sql, values)
                connection.commit()
        else:
            if text_html:
                combined_text = remove_text_and_words(text_html)
            else:
                combined_text = remove_text_and_words(text)

        retries = 0
        while retries < 3:
            try:
                completion = openai.ChatCompletion.create(
                # gpt-4-0613
                # gpt-3.5-turbo
                model="gpt-3.5-turbo",
                messages = [{"role": "user", "content": f" As an expert text analyst, analyze the provided text and extract specific details related to different products. The details that need to be extracted include (if any of detail has comma in it switch it to dot): the Product Name, Weight/Volume, Type of Packaging, Price , Type of Price (such as Per Case, Box, or Pcs), and Incoterms. Please ensure that the extracted information is formatted as a CSV file, with the following columns:'Product Name','Weight/Volume','Type of Packaging','Price','Type of Price','Incoterms'. The extracted details should be comprehensive for each product mentioned  in the text and if any of the columns are missing information, it will be marked as 'none'. The provided text is: {combined_text}."}],
                max_tokens = 1000,
                temperature = 0.8)
                # Extract the required information from the completion
                response_data = extract_info_from_ai_completion(completion)
                time.sleep(20)
                if response_data:
                    #leave while because we got data
                    break

                retries += 1 #dont ask me why this exist,but without it it doesnt works for some reason, base of programing be like :P

                # Pause for 20 seconds (Api has limit for requsts in a minute,if we procces more then 3 mail it give us api error)
            except Exception:
                retries += 1
                time.sleep(20)

        if response_data:
            # Add the response to the list
            for response in response_data:
                combined_list = response + [sender,subject,date]
                ai_responses.append(combined_list)

        else:
            send_email(subject, sender, date, EMAIL)
            sql = "INSERT INTO offers (product,weight, pack_type, price, price_type, incoterm,sender,subject,date) VALUES (%s, %s, %s, %s,%s,%s,%s,%s,%s)"
            product = 'Too big offer check manually'
            weight = 'none'
            pack_type = 'none'
            price = 'none'
            price_type = 'none'
            incoterm = 'none'
            values = (product,weight, pack_type, price, price_type, incoterm,sender,subject,date)
            cursor.execute(sql, values)
            connection.commit()


            for attachment in mail.attachments:
                filename = attachment['filename']
                if filename.endswith(('.xls', '.xlsx', '.pdf', '.csv')):
                    sql = "INSERT INTO offers (product,weight, pack_type, price, price_type, incoterm,sender,subject,date) VALUES (%s, %s, %s, %s,%s,%s,%s,%s,%s)"
                    product = filename
                    weight = 'none'
                    pack_type = 'none'
                    price = 'none'
                    price_type = 'none'
                    incoterm = 'none'
                    values = (product,weight, pack_type, price, price_type, incoterm,sender,subject,date)
                    cursor.execute(sql, values)
            connection.commit()

    save_ai_responses()  # Save AI responses after all emails are processed


def count_word_occurrences(text, word):
    """" Count how much words Product in string"""
    words = text.split()
    count = 0
    for w in words:
        if w.lower() == word.lower():
            count += 1
    return count


def check_word_repetition(text, word):
    """ Check if words Prouct more then 2"""
    count = count_word_occurrences(text, word)
    return count >= 2


def remove_last_comma(string):
    """" Remove last comma for price :)"""
    if string.endswith(","):
        string = string[:-1]
    return string


def extract_info_from_ai_completion(completion):
    """Function to extract the required information from the AI completion"""
    response = completion.choices[0].message['content']
    logging.debug('Chat answer is %s', response)

    last_price = 'none'
    last_incoterm = 'none'
    extracted_info = []
    response_io = io.StringIO(response)  # Converts string to a file-like object for csv.reader
    csv_reader = csv.reader(response_io, delimiter=',')

    next(csv_reader, None)  # Skips the header row

    for row in csv_reader:
        try:
            # Make sure the row has 6 elements, filling in "none" for any missing values
            row += ["none"] * (6 - len(row))

            product_name = row[0].strip()
            product_name = clear_string_for_db(product_name)

            weight = row[1].strip()
            weight = clear_string_for_db(weight)

            pack_type = row[2].strip()
            pack_type = clear_string_for_db(pack_type)

            price = row[3].strip()
            price = clear_string_for_db(price)
            if price.lower() == 'none':
                price = last_price  # Use last_price if current price is 'none' or 'None'
            else:
                last_price = price  # Update last_price with the current valid price
            price = remove_last_comma(price)
            price = price.replace(".", ",")


            price_type = row[4].strip()
            price_type = clear_string_for_db(price_type)

            incoterm = row[5].strip()
            if incoterm.lower() == 'none':
                incoterm = last_incoterm  # Use last_price if current price is 'none' or 'None'
            else:
                last_incoterm = incoterm  # Update last_price with the current valid price
            incoterm = clear_string_for_db(incoterm)


            extracted_info.append([product_name, weight, pack_type, price, price_type, incoterm])

        except Exception as e:
            logging.error(f"Error while extracting information from product: {e}. Skipping this product.")
            continue
    return extracted_info



def save_ai_responses():
    """Saves AI responses to the database"""

    insert_sql = """INSERT INTO offers (product, weight, pack_type, price, price_type,
                    incoterm, sender, subject, date) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""


    for data in ai_responses:
        product, weight, pack_type, price, price_type, incoterm, sender, subject, date = data

        # If not, insert the new data
        values = (product, weight, pack_type, price, price_type, incoterm, sender, subject, date)
        cursor.execute(insert_sql, values)

    connection.commit()

    print("AI responses saved to the database")



# Run the program
# Setup logging
logging.basicConfig(filename='email_processing.log', level=logging.DEBUG)

username = EMAIL  #your mail here
password = PASSWORD  #your password for apps here
parse_emails(username, password)